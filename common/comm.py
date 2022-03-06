from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import inspect
import logging
import openpyxl

class comman_class:

    def invoke_driver(self, url):
        driverservice = Service(ChromeDriverManager().install())
        d = webdriver.Chrome(service=driverservice)
        d.maximize_window()
        d.get(url)
        return d


    def close_driver(self, driver):
        driver.quit()


    def getLogger(self):
        logfile = "logfile.log"
        loggerName = inspect.stack()[1][3]
        logger = logging.getLogger(loggerName)
        fileHandler = logging.FileHandler(logfile)
        formatter = logging.Formatter("%(asctime)s :%(levelname)s : %(name)s :%(message)s")
        fileHandler.setFormatter(formatter)

        logger.addHandler(fileHandler)  # filehandler object

        logger.setLevel(logging.DEBUG)
        return logger


    def getTestData(self, test_case_name):
        test_data = "test_data.xlsx"
        Dict = {}
        book = openpyxl.load_workbook(test_data)
        sheet = book.active

        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return Dict

